from openpyxl import workbook
wb = workbook()
ws = wb.create_sheet()  #새로운 시트 기본 이름으로 생성
ws = wb.acitve # 현재 활성화된 시트를 가져옴

ws.title = "시트이름" # 시트이름 변경
ws.sheet_properties.tabColor = "rgb 값" #탭 색상 변경

ws1 = wb.create_sheet("시트네임1") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("시트네임2",2) #2번째 인덱스에 시트 생성


print(wb.sheetnames) # 모든 시트 이름 확인
 
# 데이터 넣어보기
ws["A1"] = 1
print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) #A1 셀의 '값' 출력
# row = 1, 2, 3..  #column = A, B, C..
ws.cell(row=1 , column=1).value ==  ws["A1"].value
ws.cell(row=1 , column=3, value=10) # ws["C1"].value = 10

# 반복물 이용해 랜덤 숫자 채워보기
from random import *
for x in range (1,11):
    for y in range (1,11):
        ws.cell (column=x, row=y, value=randint(0,100))
        
# 파일 불러오기
from openpyxl import load_workbook
wb = load_workbook("불러올 파일")
ws = wb.active #워크시트는 현재 활성화된 워크시트 부른다

# cell 갯수 모를 때
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell (row = x,column = y,end =" "))

#한줄씩 넣기

ws.append(["a","b","c"])

col_B = ws["b"] # b column 만 가져오기
for cell in col_B:
    print(cell.value)
    
col_range = ws["b:c"] #컬럼 이어서 가져오기
for cols in col_range:
    for coll in cols:
        print(cell.value)
    
# row 기준
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")

# xy 분류해서 작업하기
from openpyxl.utils.cell import coordinate_from_string
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate)
        print(xy[0], end=" ") #A
        print(xy[1], end=" ") #1
        
#전체 rows  
for row in tuple(ws.rows):
    print(row[2].value)
#전체 column
for column in tuple(ws.columns):
    print(column[0].value)