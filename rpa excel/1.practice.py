from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
import time

wb =  Workbook()
ws1 = wb.active 
ws1.title = "연습1"
ws2 = wb.create_sheet("연습2")

ws1.append(["1","2","3"])

print(wb.sheetnames)
wb.save("Kream.xlsx")
wb.close()
