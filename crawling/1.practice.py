from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait #엘리먼트 나올때까지 혹은 최대 지정시간까지 기다림 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

##
from openpyxl import Workbook
from openpyxl import load_workbook
import time
 
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)  # to keep browser open
driver = webdriver.Chrome(service=service, options=options)


wb = load_workbook("Kream.xlsx")

if '데이터' in wb.sheetnames:
    ws1 = wb["데이터"]
else:
    ws1 = wb.active
    ws1.title = "데이터"


record_date = []               # 날짜 모델 사이즈 가격
record_model_num = []
for date in ws1['A']:
    record_date.append(date.value)
for number in ws1['B']:
    record_model_num.append(number.value)

url = "https://kream.co.kr/login"
driver.get(url)
driver.maximize_window()

driver.find_element(By.XPATH,"//*[@id='wrap']/div[2]/div/div/div[1]/div/input").send_keys("dbsqa@naver.com")
driver.find_element(By.XPATH,"//*[@id='wrap']/div[2]/div/div/div[2]/div/input").send_keys("Xe0hJd#U")
login = driver.find_element(By.XPATH,'//*[@id="wrap"]/div[2]/div/div/div[3]/a').click()
time.sleep(2)
driver.get("https://kream.co.kr/my/buying")
time.sleep(2)
driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[2]/div[4]/a/dl').click()
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[3]/div[1]/ul/li[3]/a').click()
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[3]/div[2]/div[2]/button').click()
time.sleep(2)

item_list = driver.find_elements(By.CLASS_NAME, "list_item_img")
state_list = driver.find_elements(By.CLASS_NAME, "last_title display_paragraph")

buy_datas = [] #driver.find_element(By.CLASS_NAME,"price_text")
sizes = [] #driver.find_element(By.CLASS_NAME,"product_option")
model_names = []  #driver.find_element(By.CLASS_NAME,"product_title")
model_numbers= [] #driver.find_element(By.CLASS_NAME,"product_description").text
prices = []  #driver.find_element(By.CLASS_NAME,"amount")
flag = 0
page = 0
cnt = 0 
while True:
    item_list = driver.find_elements(By.CLASS_NAME, "list_item_img")
    state_list = driver.find_elements(By.CSS_SELECTOR, 'p.last_title.display_paragraph')
    if cnt >=10:
        break
    for i in range(0,len(item_list)):
        cnt+=1
        print("i=",i, "itam_list 갯수=", len(item_list))
        item_list = driver.find_elements(By.CLASS_NAME, "list_item_img")
        state_list = driver.find_elements(By.CSS_SELECTOR, 'p.last_title.display_paragraph')
        
        if state_list[i].text != "배송완료":
            continue
        item_list[i].click()
        time.sleep(1)
        
        buy_data = driver.find_element(By.CLASS_NAME,"price_text").text
        buy_datas.append(buy_datas)
        
        model_name = driver.find_element(By.CLASS_NAME,"product_title").text
        model_names.append(model_name)
        
        model_number = driver.find_element(By.CLASS_NAME,"product_description").text
        model_numbers.append(model_number)
        
        size = driver.find_element(By.CLASS_NAME,"product_option").text
        sizes.append(size)
    
        price = driver.find_element(By.CLASS_NAME,"amount").text
        prices.append(price)
        
        print(buy_data,model_name,size,price,model_number)
        
        if buy_data in record_date:
            if model_name in record_model_num:
                flag = 1
                break
            
        ws1.append([buy_data, model_name, size, price, model_number])
        driver.back()
        time.sleep(1)
    if flag == 1:
        break
    if len(item_list) != 10:
        break
    

    # page_list = driver.find_elements(By.CLASS_NAME,"btn_page")
    # page+=1
    
    # if page % 5 == 0:
    #     driver.find_element(By.CLASS_NAME, "arr-page-next icon sprite-icons").click()
    # else:
    #     page_list[page%5].click()
    
    driver.execute_script('window.scrollTo(0,0)')

print(len(ws1["A"]))
wb.save(filename='Kream.xlsx')
driver.quit()