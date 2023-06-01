from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait #엘리먼트 나올때까지 혹은 최대 지정시간까지 기다림 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

##
from openpyxl import Workbook
from openpyxl import load_workbook
import time
 
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)  # to keep browser open
driver = webdriver.Chrome(service=service, options=options)


wb =  load_workbook('솔드아웃 구매내역 스크래핑_원본.xlsx')
print(wb.sheetnames)

if '원본데이터' in wb.sheetnames:
    ws1 = wb['원본데이터']
else:
    ws1 = wb.active
    ws1.title = "원본데이터"

list_purchase_date_before = []
list_serial_num_before = []

for data1 in ws1['A']:
    list_purchase_date_before.append(data1.value)
for data2 in ws1['B']:
    list_serial_num_before.append(data2.value)


url = "https://www.soldout.co.kr/member/login"

driver.get(url)
driver.maximize_window()

driver.find_element(By.CSS_SELECTOR,"input[placeholder='아이디']").send_keys("rladnr2222")
driver.find_element(By.CSS_SELECTOR,"input[placeholder='비밀번호']").send_keys("di1sj2rn3!")

driver.find_element(By.XPATH, '//*[@id="__layout"]/div/div[2]/div/form/button').click()
#driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div/div[3]/a').click()
time.sleep(2)
driver.get("https://www.soldout.co.kr/my/buy")
time.sleep(5)
driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/section/div/div[2]/div/div[2]/div[1]/div[1]/ul/li[6]/span[1]').click()
time.sleep(1)


item_list = driver.find_elements(By.CLASS_NAME, 'table-items')
print(len(item_list))

purchase_date = []
serial_num = []
size = []
price = []
page = 0
flag = 0


while True:
    item_list = driver.find_elements(By.CLASS_NAME, 'table-items')
    for i in range(0, len(item_list)):
        print("i=",i, "itam_list 갯수=", len(item_list))
        item_list = driver.find_elements(By.CLASS_NAME, 'table-items')
        item_list[i].click()
        time.sleep(2)
        purchase_date_temp = driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/section/div/div[2]/div/div[2]/div[2]/div[3]/div/div/div[2]/span[2]').text
        purchase_date.append(purchase_date_temp)

        size_temp = driver.find_element(By.CLASS_NAME,'size').text
        size.append(size_temp)

        serial_num_temp = driver.find_element(By.CLASS_NAME,'ellipsis').text
        serial_num.append(serial_num_temp)

        price_temp = driver.find_element(By.CSS_SELECTOR,'span.result.bold').text.strip('원')
        price.append(price_temp)
        if purchase_date_temp in list_purchase_date_before:
            if serial_num_temp in list_serial_num_before:
                print(purchase_date_temp, serial_num_temp, size_temp, price_temp)
                print("여기까지 저번에 했음")
                flag = 1
                break
        ws1.append([purchase_date_temp, serial_num_temp, size_temp, price_temp])
        print(purchase_date_temp, serial_num_temp, size_temp, price_temp)
        driver.find_element(By.XPATH, '//*[@id="__layout"]/div/div[2]/section/div/div[2]/div/div[2]/div[3]/button').click()
        time.sleep(1)
    if flag == 1:
        break
    #wb.save(filename='크림 구매내역 스크래핑.xlsx')
    if len(item_list) != 10:
        break

    page = page + 1
    print("page = ", page)
    driver.find_element(By.CLASS_NAME, 'next').click()
   
    time.sleep(1)
 #   driver.execute_script('window.scrollTo(0,0)')
 #   time.sleep(1)

driver.quit()

row_max = ws1.max_row
ws1["F1"] = "가격(숫자)"
ws1["E1"] = "품번(사이즈)"
#중복 제거 리스트 선언
serial_size = []
sort_serial_size = []
for row in range(2, row_max+1):
    ws1["F"+str(row)].value = "=VALUE(D"+str(row)+")"
    ws1["E"+str(row)].value = ws1["B"+str(row)].value + "(" + ws1["C"+str(row)].value +")"
    if not ws1["E"+str(row)].value in serial_size:
        serial_size.append(ws1["E"+str(row)].value)
ws1["G1"] = "품번(사이즈) 중복제거"
sort_serial_size = sorted(serial_size)

for i in range(0, len(serial_size)):
    ws1["G"+str(i+2)].value = serial_size[i]

ws1["H1"] = "평균구입가격"
ws1["I1"] = "구매갯수"
for row in range(2, len(serial_size)+2 ):
    ws1["H"+str(row)] = "=ROUNDDOWN(AVERAGEIF(E:E,G"+str(row)+",F:F),0)"
    ws1["I"+str(row)] = "=COUNTIF(E:E,G"+str(row)+")"

if not '평균구매가격' in wb.sheetnames:
    wb.create_sheet('평균구매가격',0)
ws2 = wb['평균구매가격']

for row in range(1, len(serial_size)+2):
    ws2["A"+str(row)] = "=원본데이터!G"+str(row)
    ws2["B"+str(row)] = "=원본데이터!H"+str(row)
    ws2["C"+str(row)] = "=원본데이터!I"+str(row)

wb.save(filename='솔드아웃 구매내역 스크래핑_원본.xlsx')





