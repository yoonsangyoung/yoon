from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait #엘리먼트 나올때까지 혹은 최대 지정시간까지 기다림 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

##
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

import urllib.request
import time


from collections import Counter
import re
import os

service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)  # to keep browser open
driver = webdriver.Chrome(service=service, options=options)

wb = load_workbook("Kream.xlsx")
ws1 = wb["원본데이터"]


model_list = [cell.value for cell in ws1["B"]]
dic_model_list = Counter(model_list)
serach_list = [k for k,v in dic_model_list.items() if v>=100]

print(serach_list,len(serach_list))

# ##

url = "https://kream.co.kr/login"
driver.get(url)
driver.maximize_window()

driver.find_element(By.XPATH,"//*[@id='wrap']/div[2]/div/div/div[1]/div/input").send_keys("dbsqa@naver.com")
driver.find_element(By.XPATH,"//*[@id='wrap']/div[2]/div/div/div[2]/div/input").send_keys("Xe0hJd#U")
login = driver.find_element(By.XPATH,'//*[@id="wrap"]/div[2]/div/div/div[3]/a').click()
time.sleep(2)

ws2 = wb["시세"]
cnt = 0

for data  in serach_list:
    cnt+=1
    url = "https://kream.co.kr/search?keyword="+data
    print(url,"//",cnt)
    driver.get(url)
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR,"div.product").click()
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR,"strong.title").click()
    time.sleep(3)
    price_records = driver.find_elements(By.CLASS_NAME,"select_item")
    
    
    # Capture = driver.get_screenshot_as_png()
    # Capture = PILImage.open(BytesIO(Capture))  
    # crop = Capture.crop((700, 250, 2300, 1600))
    # crop.save(data+'.png') 
    
    # img = XLImage(data+'.png')
    # ws2.add_image(img, 'A'+str(cnt))
    # wb.save(filename='Kream.xlsx')
    
    # Capture_position = driver.find_element(By.CSS_SELECTOR,'div.buy_before_check')
    # location = Capture_position.location
    # how_size = Capture_position.size
    # left = location['x']
    # right = left + how_size['width']
    # top = location['y']
    # bottom = top + how_size['height']
    # crop2 = Capture.crop((left,top,right,bottom))
    # crop2.save(data+'.png')
    # print(location)
    # print(how_size)

    # 신발 로고 이미지 가져오기
    # img_url = driver.find_element(By.CSS_SELECTOR, "img[src]").get_attribute("src")
    # folder_path = "/Users/Yoon/Desktop/python/temp_img"
    # file_name = "1.png"
    # urllib.request.urlretrieve(url, folder_path + file_name)
    # img = Image(folder_path + file_name)
    
    
    for price_record in price_records:
        name = data
        size = price_record.find_element(By.CSS_SELECTOR, "span.size").text
        size = re.findall('\d+', size)[0]
        price = price_record.find_element(By.CSS_SELECTOR, "span.price").text
        if price == "구매입찰":
            continue
        print(size,price)
        
        for col_num in range(2, 24):
            cell_value = ws2.cell(row=1, column=col_num).value
            if str(size) == str(cell_value):
                ws2.cell(row=cnt+1, column=2, value=name)
                ws2.cell(row=cnt+1, column=col_num, value=price)
                print("저장")
                break
wb.save(filename='Kream.xlsx')

