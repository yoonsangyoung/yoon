from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait #엘리먼트 나올때까지 혹은 최대 지정시간까지 기다림 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

##
from openpyxl import Workbook
from openpyxl import load_workbook
import time
 
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)  # to keep browser open
driver = webdriver.Chrome(service=service, options=options)


wb =  load_workbook('크림 구매내역 스크래핑_원본.xlsx')

if '원본데이터' in wb.sheetnames:
    ws1 = wb['원본데이터']
else:
    ws1 = wb.active
    ws1.title = "원본데이터"

# list_purchase_date_before = []
list_serial_num_before = []
order_num_before = []

# for data1 in ws1['A']:
#     list_purchase_date_before.append(data1.value)
for data2 in ws1['B']:
    list_serial_num_before.append(data2.value)
for data3 in ws1['J']:
    order_num_before.append(data3.value)

url = "https://kream.co.kr/login"

driver.get(url)
driver.maximize_window()

driver.find_element(By.CSS_SELECTOR,"input[type='email']").send_keys("akilless_@naver.com")
driver.find_element(By.CSS_SELECTOR,"input[type='password']").send_keys("di1sj2rn3!")

login_button = driver.find_elements(By.LINK_TEXT, "로그인")
login_button[1].click()
#driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[1]/div/div[3]/a').click()
time.sleep(2)
driver.get("https://kream.co.kr/my/buying")
time.sleep(3)
driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[2]/div[4]/a/dl').click()

time.sleep(3)

driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[3]/div[1]/ul/li[3]/a').click()
time.sleep(1)
driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[3]/div[2]/div[2]/button').click()
time.sleep(2)

item_list = driver.find_elements(By.CLASS_NAME, 'list_item_img')
#print(len(item_list))
state_list = driver.find_elements(By.CSS_SELECTOR, 'p.last_title.display_paragraph')
#print(len(state_list))

purchase_date = []
serial_num = []
size = []
price = []
order_num = []
page = 0
flag = 0


while True:
    item_list = driver.find_elements(By.CLASS_NAME, 'list_item_img')
    state_list = driver.find_elements(By.CSS_SELECTOR, 'p.last_title.display_paragraph')
    for i in range(0, len(item_list)):
        print("i=",i, "itam_list 갯수=", len(item_list))
        item_list = driver.find_elements(By.CLASS_NAME, 'list_item_img')
        state_list = driver.find_elements(By.CSS_SELECTOR, 'p.last_title.display_paragraph')
        purchase_date = driver.find_elements(By.CLASS_NAME, 'secondary_title.display_paragraph')
        if state_list[i].text != "배송완료":
            continue
        
        day = str(purchase_date[i].text)
        item_list[i].click()
        time.sleep(1)
        # purchase_date_temp = driver.find_element(By.XPATH,'//*[@id="__layout"]/div/div[2]/div[2]/div/div[4]/div[2]/div[1]/div[3]/dl/dd').text
        # purchase_date.append(purchase_date_temp)
        # print(purchase_date_temp[4].text)
        

        size_temp = driver.find_element(By.CLASS_NAME,'product_option').text
        size.append(size_temp)

        serial_num_temp = driver.find_element(By.CLASS_NAME,'product_description').text
        serial_num.append(serial_num_temp)

        price_temp = driver.find_element(By.CLASS_NAME,'amount').text
        price.append(price_temp)
        
        order_num_temp = driver.find_element(By.CLASS_NAME, 'order_number').text
        order_num.append(order_num_temp)
        
        if order_num_temp in order_num_before:
            print(order_num_temp,serial_num_temp, size_temp, price_temp)
            print("여기까지 저번에 했음")
            flag = 1
            break
        ws1.append([day] + [serial_num_temp, size_temp, price_temp] + [None]*5 + [order_num_temp])
        
        print(day, serial_num_temp, size_temp, price_temp, order_num_temp)
        driver.find_element(By.LINK_TEXT, "목록보기").click()
        
        time.sleep(1)
    if flag == 1:
        break
    #wb.save(filename='크림 구매내역 스크래핑.xlsx')
    if len(item_list) != 10:
        break

    page = page + 1
    print("page = ", page)
    page_list = driver.find_elements(By.CLASS_NAME,'btn_page')
    if page%5 == 0:
        driver.find_element(By.CLASS_NAME,'arr-page-next').click()
    else:
        page_list[page%5].click()
    time.sleep(1)
    driver.execute_script('window.scrollTo(0,0)')
    time.sleep(1)

driver.quit()

row_max = ws1.max_row
ws1["F1"] = "가격(숫자)"
ws1["E1"] = "품번(사이즈)"
#중복 제거 리스트 선언
serial_size = []
sort_serial_size = []
for row in range(2, row_max+1):
    ws1["F"+str(row)].value = "=VALUE(D"+str(row)+")"
    ws1["E"+str(row)].value = ws1["B"+str(row)].value + "(" + ws1["C"+str(row)].value +")"
    if not ws1["E"+str(row)].value in serial_size:
        serial_size.append(ws1["E"+str(row)].value)
ws1["G1"] = "품번(사이즈) 중복제거"
sort_serial_size = sorted(serial_size)

for i in range(0, len(serial_size)):
    ws1["G"+str(i+2)].value = serial_size[i]

ws1["H1"] = "평균구입가격"
ws1["I1"] = "구매갯수"
for row in range(2, len(serial_size)+2 ):
    ws1["H"+str(row)] = "=ROUNDDOWN(AVERAGEIF(E:E,G"+str(row)+",F:F),0)"
    ws1["I"+str(row)] = "=COUNTIF(E:E,G"+str(row)+")"

if not '평균구매가격' in wb.sheetnames:
    wb.create_sheet('평균구매가격',0)
ws2 = wb['평균구매가격']

for row in range(1, len(serial_size)+2):
    ws2["A"+str(row)] = "=원본데이터!G"+str(row)
    ws2["B"+str(row)] = "=원본데이터!H"+str(row)
    ws2["C"+str(row)] = "=원본데이터!I"+str(row)

wb.save(filename='크림 구매내역 스크래핑_원본.xlsx')





